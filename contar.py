from exchangelib import Credentials, Account, Folder, FileAttachment, ItemAttachment, errors
from EWS_Conection import Cuenta
from sucursales import lista_sucursales, list_back, list_back2
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import datetime

wb = Workbook()
i = 2
errorList = []

# grab the active worksheet
ws = wb.active

ws['A1'] = 'ID'
ws['B1'] = 'Sucursal'
ws['C1'] = 'Correo'
ws['D1'] = 'Cantidad Inbox'
ws['E1'] = 'Estado'
ws['F1'] = 'RPA-Tramitados'
ws['G1'] = 'RPA-Rechazados'


# Select the row and apply the background color to each cell
row_number = 1
fill = PatternFill(start_color='FFFFFF', end_color='0000FF', fill_type='mediumGray')
for cell in ws[row_number]:
    cell.fill = fill    

for cuentas in list_back2:
    # Cuenta la cantidad de correo en la casilla y captura el error si hay.
    ws[f'A{i}'] = cuentas["id"]
    ws[f'B{i}'] = cuentas["sucursal"]
    ws[f'C{i}'] = cuentas["mail"]
    cell_E = ws[f'E{i}']
    try:
        account = Cuenta(cuentas["mail"])
        print(account.inbox.all().count())
        ws[f'D{i}'] = account.inbox.all().count()
        
        if cuentas["RPA"]:
            folder_t = account.inbox / "RPA-Tramitados"
            folder_r = account.inbox / "RPA-Rechazados"
            ws[f'F{i}'] = folder_t.all().count()                   
            ws[f'G{i}'] = folder_r.all().count()
        
        if account.inbox.all().count() > 200:
            # Select the cell and apply the background color
            fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            cell_E.fill = fill
            cell_E.value = 'Verificar'
        else:
            fill = PatternFill(start_color='77AC42', end_color='77AC42', fill_type='solid')
            cell_E.fill = fill
            cell_E.value = 'OK'
     
    except errors.ErrorFolderNotFound:
            fill = PatternFill(start_color='77AC42', end_color='77AC42', fill_type='solid')
            cell_E.fill = fill
            cell_E.value = 'OK'
    except Exception as e:
            errorList.append(cuentas)
            fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            cell_E.fill = fill
            cell_E.value = 'Falla conexion'
            print(cuentas)
            print(e)
        
    i += 1


today = datetime.date.today().strftime("%Y-%m-%d")
print(errorList)
wb.save(f'./informes/SUCURSALES_INBOX_&_RPA_{today}.xlsx') 